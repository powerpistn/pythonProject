import requests
import json
from openpyxl import load_workbook
import time
# 加载Excel文件
workbook = load_workbook('/test.xlsx')
# 选择第一个工作表
worksheet = workbook.worksheets[0]
url = 'http://localhost:5000/generate'
# 遍历每一行
for row_number, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
    # 获取每行第5列的数据
    query = row[4].value #从0开始计数
    print("query is ",query)
    start_time = time.time()
    response = requests.post(url, json={'query': query})
    generated_text = response.json()
    print(generated_text)
    end_time = time.time()
    generated_text = generated_text["generated_text"]
    answer = generated_text.split('\n')[1] if "\n" in generated_text else generated_text  #只是为了处理generated_text和操作xlsx无关
    response_time = (end_time-start_time)*1000
    print("生成的答案:", answer)
    print("time is ",response_time)
    # 写入，将得到的结果写入每行的第8列，时间写入每行的第9列
    cell = worksheet.cell(row=row_number, column=8) 
    cell.value = answer
    cell = worksheet.cell(row=row_number, column=9)##这种方式列的索引从1开始
    cell.value = response_time
workbook.save('/test.xlsx')